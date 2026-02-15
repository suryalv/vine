"""
Generate sample Excel (.xlsx) files for UW Companion testing.
Creates realistic underwriting spreadsheet data.

Usage:
    python generate_xlsx.py
"""

from openpyxl import Workbook


def generate_statement_of_values():
    """Generate a Statement of Values spreadsheet with property and loss data."""
    wb = Workbook()

    # Sheet 1: Property Schedule
    ws1 = wb.active
    ws1.title = "Property Schedule"

    headers = [
        "Location #", "Address", "City", "State",
        "Building Value", "Contents Value", "BI Value",
        "Coverage Limit", "Construction Type", "Year Built",
        "Occupancy", "Protection Class",
    ]
    ws1.append(headers)

    locations = [
        [1, "100 Industrial Blvd", "Houston", "TX", 5200000, 1800000, 2500000, 9500000, "Fire Resistive", 2015, "Manufacturing", 3],
        [2, "250 Commerce Dr", "Dallas", "TX", 3800000, 950000, 1200000, 5950000, "Masonry Non-Combustible", 2008, "Warehouse", 4],
        [3, "75 Tech Park Way", "Austin", "TX", 8500000, 3200000, 4100000, 15800000, "Fire Resistive", 2020, "Office", 2],
        [4, "1200 Harbor Rd", "Galveston", "TX", 2100000, 680000, 900000, 3680000, "Frame", 1995, "Retail", 6],
        [5, "500 Energy Center", "Midland", "TX", 12000000, 5500000, 7200000, 24700000, "Fire Resistive", 2018, "Petro-Chemical", 3],
        [6, "333 Market St", "San Antonio", "TX", 4200000, 1100000, 1500000, 6800000, "Joisted Masonry", 2012, "Mixed Use", 4],
        [7, "88 Airport Blvd", "El Paso", "TX", 6700000, 2800000, 3200000, 12700000, "Non-Combustible", 2017, "Distribution Center", 3],
        [8, "1500 Campus Dr", "Fort Worth", "TX", 9100000, 4000000, 5600000, 18700000, "Fire Resistive", 2022, "Corporate HQ", 2],
    ]

    for loc in locations:
        ws1.append(loc)

    # Sheet 2: Loss History
    ws2 = wb.create_sheet("Loss History")
    loss_headers = [
        "Claim #", "Date of Loss", "Location #", "Description",
        "Incurred Amount", "Paid Amount", "Status", "Cause of Loss",
    ]
    ws2.append(loss_headers)

    losses = [
        ["CLM-2023-001", "2023-03-15", 1, "Wind damage to roof section", 125000, 118500, "Closed", "Windstorm"],
        ["CLM-2023-002", "2023-07-22", 4, "Water damage from pipe burst", 45000, 42000, "Closed", "Water Damage"],
        ["CLM-2024-001", "2024-01-10", 2, "Forklift collision with rack system", 78000, 78000, "Closed", "Equipment Damage"],
        ["CLM-2024-002", "2024-04-05", 5, "Lightning strike to electrical panel", 210000, 195000, "Closed", "Lightning"],
        ["CLM-2024-003", "2024-09-12", 1, "Hurricane wind damage", 350000, 280000, "Open", "Named Storm"],
        ["CLM-2025-001", "2025-02-01", 3, "Theft of IT equipment", 92000, 0, "Open", "Theft"],
    ]

    for loss in losses:
        ws2.append(loss)

    # Sheet 3: Coverage Summary
    ws3 = wb.create_sheet("Coverage Summary")
    cov_headers = [
        "Coverage", "Limit", "Deductible", "Coinsurance",
        "Valuation", "Premium", "Notes",
    ]
    ws3.append(cov_headers)

    coverages = [
        ["Building", 51600000, 25000, "80%", "Replacement Cost", 128000, "All locations combined"],
        ["Contents", 20030000, 10000, "80%", "Replacement Cost", 48000, "All locations combined"],
        ["Business Income", 26200000, "72 hours", "50%", "Actual Loss Sustained", 65000, "12-month period of indemnity"],
        ["Equipment Breakdown", 51600000, 10000, "N/A", "Replacement Cost", 15000, "Included in property"],
        ["Flood", 5000000, 100000, "N/A", "Actual Cash Value", 22000, "Sub-limit per location"],
        ["Earthquake", 10000000, "5% of TIV", "N/A", "Replacement Cost", 18000, "Annual aggregate deductible"],
        ["Wind/Hail", 51600000, "2% of TIV", "N/A", "Replacement Cost", 35000, "Per-occurrence deductible"],
        ["Ordinance or Law", 5000000, 25000, "N/A", "N/A", 8000, "Coverage A, B, C combined"],
    ]

    for cov in coverages:
        ws3.append(cov)

    output_path = "Statement_of_Values_Spreadsheet.xlsx"
    wb.save(output_path)
    print(f"Generated: {output_path}")
    return output_path


if __name__ == "__main__":
    generate_statement_of_values()
